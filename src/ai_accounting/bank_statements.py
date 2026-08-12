"""Pure bank-statement parsing, hashing, and reconciliation calculations.

The functions in this module do not open paths, access a database, or write
audit state.  A caller loads a statement once and passes the same immutable
``bytes`` object that is hashed and parsed here.
"""

from __future__ import annotations

import csv
import hashlib
import json
import uuid
import zipfile
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from .bank_statement_schemas import (
    MAX_SIGNED_BIGINT,
    MIN_SIGNED_BIGINT,
    BankReconciliationPreview,
    BankReconciliationSystemFacts,
    BankStatementFileFormat,
    BankStatementImportEvidenceFact,
    BankStatementImportPreview,
    BankStatementImportRowSystemFact,
    BankStatementImportSystemFacts,
    BankStatementInformationRequirement,
    BankStatementIssue,
    BankStatementNormalizedRow,
    BankStatementParseStatus,
    BankStatementPreviewRow,
    BankStatementPreviewStatus,
    BankStatementTransactionSnapshot,
    MissingExternalIdResolution,
    ParsedBankStatement,
    PreviewBankReconciliationRequest,
    PreviewBankStatementImportRequest,
)

_CANONICAL_HASH_VERSION = "bank-statement-preview-v1"
_CHINA_TIME_ZONE = ZoneInfo("Asia/Shanghai")
MAX_STATEMENT_ROWS = 100_000
MAX_STATEMENT_COLUMNS = 200
MAX_SOURCE_CELL_CHARACTERS = 10_000
MAX_STATEMENT_BYTES = 20 * 1024 * 1024
MAX_XLSX_ARCHIVE_MEMBERS = 2_000
MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_TOTAL_UNCOMPRESSED_BYTES = 200 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 1_000


class _InputError(ValueError):
    def __init__(self, code: str, *, field_path: str | None = None) -> None:
        self.code = code
        self.field_path = field_path
        super().__init__(code)


class _RowError(ValueError):
    def __init__(self, code: str, field_path: str) -> None:
        self.code = code
        self.field_path = field_path
        super().__init__(code)


def canonical_sha256(payload: Any) -> str:
    """Hash JSON-compatible facts without accepting binary floating point."""

    canonical = json.dumps(
        _canonical_value(payload),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def canonical_json(payload: Any) -> str:
    """Serialize the exact JSON form whose SHA-256 is used by bank workflows."""

    return json.dumps(
        _canonical_value(payload),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def parse_bank_statement_bytes(
    request: PreviewBankStatementImportRequest,
    statement_bytes: bytes,
) -> ParsedBankStatement:
    """Parse one byte string once; this result is not a formal import preview."""

    source_sha256 = hashlib.sha256(statement_bytes).hexdigest()
    parser_fingerprint = canonical_sha256(_parser_binding_payload(request, source_sha256))
    if len(statement_bytes) > MAX_STATEMENT_BYTES:
        return ParsedBankStatement(
            status=BankStatementParseStatus.REJECTED,
            source_sha256=source_sha256,
            parser_request_fingerprint_sha256=parser_fingerprint,
            errors=[BankStatementIssue(code="BANK_STATEMENT_FILE_LIMIT_EXCEEDED")],
            trace=[
                {
                    "stage": "statement_parse_rejected",
                    "code": "BANK_STATEMENT_FILE_LIMIT_EXCEEDED",
                }
            ],
        )
    try:
        source_rows = _read_rows_once(request, statement_bytes)
    except _InputError as exc:
        return ParsedBankStatement(
            status=BankStatementParseStatus.REJECTED,
            source_sha256=source_sha256,
            parser_request_fingerprint_sha256=parser_fingerprint,
            errors=[BankStatementIssue(code=exc.code, field_path=exc.field_path)],
            trace=[{"stage": "statement_parse_rejected", "code": exc.code}],
        )

    normalized_rows: list[BankStatementNormalizedRow] = []
    row_errors: list[BankStatementIssue] = []
    for row_number, source_row in source_rows:
        try:
            normalized = _normalize_row(
                source_row,
                request.column_mapping.model_dump(exclude_none=True),
                request.date_format,
            )
        except _RowError as exc:
            row_errors.append(
                BankStatementIssue(
                    code=exc.code,
                    row_number=row_number,
                    field_path=exc.field_path,
                )
            )
            continue
        row_identity = canonical_sha256(
            {
                "version": _CANONICAL_HASH_VERSION,
                "source_sha256": source_sha256,
                "row_number": row_number,
                "row": normalized,
            }
        )
        row = BankStatementNormalizedRow(
            row_number=row_number,
            row_identity_sha256=row_identity,
            booking_date=normalized["booking_date"],
            amount_fen=normalized["amount_fen"],
            currency="CNY",
            external_id=normalized["external_id"],
            counterparty_name=normalized["counterparty_name"],
            memo=normalized["memo"],
        )
        normalized_rows.append(row)
    return ParsedBankStatement(
        status=BankStatementParseStatus.PARSED,
        source_sha256=source_sha256,
        parser_request_fingerprint_sha256=parser_fingerprint,
        rows=normalized_rows,
        errors=row_errors,
        trace=[
            {"stage": "statement_bytes_hashed", "source_sha256": source_sha256},
            {
                "stage": "statement_rows_preflighted",
                "row_count": len(source_rows),
                "valid_row_count": len(normalized_rows),
                "row_error_count": len(row_errors),
            },
        ],
        data={
            "row_count": len(source_rows),
            "valid_row_count": len(normalized_rows),
            "row_error_count": len(row_errors),
        },
    )


def preview_bank_statement_import(
    request: PreviewBankStatementImportRequest,
    parsed_statement: ParsedBankStatement,
    system_facts: BankStatementImportSystemFacts,
) -> BankStatementImportPreview:
    """Create a formal preview from parser output plus mandatory server facts."""

    if parsed_statement.status == BankStatementParseStatus.REJECTED:
        return BankStatementImportPreview(
            status=BankStatementPreviewStatus.REJECTED,
            source_sha256=parsed_statement.source_sha256,
            errors=parsed_statement.errors,
            trace=[{"stage": "statement_parse_rejected"}],
        )
    expected_parser_fingerprint = canonical_sha256(
        _parser_binding_payload(request, parsed_statement.source_sha256)
    )
    if parsed_statement.parser_request_fingerprint_sha256 != expected_parser_fingerprint:
        return _minimal_import_preview(
            parsed_statement.source_sha256,
            errors=[BankStatementIssue(code="BANK_STATEMENT_PARSED_REQUEST_MISMATCH")],
        )

    system_errors, row_facts, evidence_facts = _validate_import_system_facts(
        request,
        parsed_statement,
        system_facts,
    )
    if system_errors:
        return _minimal_import_preview(
            parsed_statement.source_sha256,
            errors=system_errors,
        )

    rows = [_preview_row(row, row_facts[row.row_identity_sha256]) for row in parsed_statement.rows]
    hard_errors: list[BankStatementIssue] = []
    for row in rows:
        _apply_existing_source_row_snapshot(
            request,
            row,
            row_facts[row.row_identity_sha256],
            hard_errors,
        )
    rows_by_identity = {row.row_identity_sha256: row for row in rows}
    missing: list[BankStatementInformationRequirement] = []
    resolution_errors = _apply_external_id_resolutions(
        request.missing_external_id_resolutions,
        rows_by_identity,
        missing,
    )
    hard_errors.extend(resolution_errors)
    resolution_by_identity = {
        item.row_identity_sha256: item for item in request.missing_external_id_resolutions
    }
    external_id_rows: dict[str, list[BankStatementPreviewRow]] = {}
    for row in rows:
        if row.external_id is not None:
            external_id_rows.setdefault(row.external_id, []).append(row)
    for duplicate_rows in external_id_rows.values():
        if len(duplicate_rows) > 1:
            for row in duplicate_rows:
                hard_errors.append(
                    BankStatementIssue(
                        code="BANK_STATEMENT_DUPLICATE_EXTERNAL_ID_IN_SOURCE",
                        row_number=row.row_number,
                        field_path="external_id",
                    )
                )

    for row in rows:
        fact = row_facts[row.row_identity_sha256]
        if not _period_projection_matches_row(row, fact):
            hard_errors.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_PERIOD_PROJECTION_MISMATCH",
                    row_number=row.row_number,
                )
            )
        if (
            fact.period.closed_at is not None
            and fact.period.closed_at.astimezone(_CHINA_TIME_ZONE).date() > system_facts.as_of_date
        ):
            hard_errors.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_PERIOD_FACTS_AS_OF_MISMATCH",
                    row_number=row.row_number,
                )
            )
        if row.booking_date > system_facts.as_of_date:
            hard_errors.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_FUTURE_BOOKING_DATE_NOT_ALLOWED",
                    row_number=row.row_number,
                    field_path="booking_date",
                )
            )
        if fact.period.status in {"not_generated", "control_disabled"}:
            missing.append(
                BankStatementInformationRequirement(
                    code=(
                        "BANK_STATEMENT_PERIOD_NOT_GENERATED"
                        if fact.period.status == "not_generated"
                        else "BANK_STATEMENT_PERIOD_CONTROL_DISABLED"
                    ),
                    fields=[f"rows.{row.row_identity_sha256}.booking_date"],
                )
            )
        _apply_existing_external_id_snapshot(request, row, fact, hard_errors)
        _validate_manual_duplicate_snapshot(
            request,
            row,
            fact,
            resolution_by_identity.get(row.row_identity_sha256),
            hard_errors,
        )

    unresolved = [row for row in rows if row.disposition == "needs_external_id_resolution"]
    if unresolved:
        missing.append(
            BankStatementInformationRequirement(
                code="BANK_STATEMENT_EXTERNAL_ID_RESOLUTION_REQUIRED",
                fields=[
                    f"missing_external_id_resolutions.{row.row_identity_sha256}"
                    for row in unresolved
                ],
            )
        )
    if parsed_statement.errors and not request.proceed_with_known_row_errors:
        missing.append(
            BankStatementInformationRequirement(
                code="BANK_STATEMENT_ROW_ERRORS_CONFIRMATION_REQUIRED",
                fields=["proceed_with_known_row_errors"],
            )
        )

    importable = [row for row in rows if row.disposition in {"ready", "manual_new"}]
    duplicates = [
        row for row in rows if row.disposition in {"stable_duplicate", "manual_duplicate"}
    ]
    if not importable and not duplicates and not missing:
        hard_errors.append(BankStatementIssue(code="BANK_STATEMENT_NO_IMPORTABLE_ROWS"))
    if hard_errors:
        return _minimal_import_preview(
            parsed_statement.source_sha256,
            errors=[*parsed_statement.errors, *hard_errors],
            rows=rows,
        )
    if missing:
        return BankStatementImportPreview(
            status=BankStatementPreviewStatus.NEEDS_INFORMATION,
            source_sha256=parsed_statement.source_sha256,
            rows=rows,
            missing_information=missing,
            errors=parsed_statement.errors,
            trace=[{"stage": "bank_statement_preview_needs_information"}],
        )

    resolution_payloads = sorted(
        (
            {
                **resolution.model_dump(mode="json"),
                "evidence_references": sorted(str(item) for item in resolution.evidence_references),
            }
            for resolution in request.missing_external_id_resolutions
        ),
        key=lambda item: (str(item["row_identity_sha256"]), int(item["row_number"])),
    )
    payload = {
        "version": _CANONICAL_HASH_VERSION,
        "command": "finance_preview_bank_statement_import",
        "request": {
            **request.model_dump(mode="json"),
            "missing_external_id_resolutions": resolution_payloads,
        },
        "parsed_statement": {
            "source_sha256": parsed_statement.source_sha256,
            "parser_request_fingerprint_sha256": (
                parsed_statement.parser_request_fingerprint_sha256
            ),
            "rows": [row.model_dump(mode="json") for row in parsed_statement.rows],
            "row_errors": [issue.model_dump(mode="json") for issue in parsed_statement.errors],
        },
        "system_facts": {
            "org_id": system_facts.org_id,
            "bank_account_code": system_facts.bank_account_code,
            "as_of_date": system_facts.as_of_date,
            "rows": [
                item.model_dump(mode="json")
                for item in sorted(
                    system_facts.rows,
                    key=lambda item: item.row_identity_sha256,
                )
            ],
            "resolution_evidence": [
                item.model_dump(mode="json")
                for item in sorted(
                    evidence_facts.values(),
                    key=lambda item: str(item.evidence_id),
                )
            ],
        },
        "preview_rows": [row.model_dump(mode="json") for row in rows],
    }
    calculation_hash = canonical_sha256(payload)
    return BankStatementImportPreview(
        status=BankStatementPreviewStatus.CALCULATED,
        calculation_hash=calculation_hash,
        source_sha256=parsed_statement.source_sha256,
        rows=rows,
        trace=[
            {"stage": "formal_system_facts_validated"},
            {"stage": "calculation_hashed", "calculation_hash": calculation_hash},
        ],
        data={
            "row_count": int(parsed_statement.data.get("row_count", len(rows))),
            "valid_row_count": len(rows),
            "planned_import_count": len(importable),
            "planned_duplicate_count": len(duplicates),
            "late_import_count": sum(row.is_late for row in importable),
            "row_error_count": len(parsed_statement.errors),
            "partial_import_expected": bool(parsed_statement.errors),
            "planned_confirm_status": ("partially_posted" if parsed_statement.errors else "posted"),
            "formal_system_facts_applied": True,
            "calculation_payload": payload,
        },
    )


def _parser_binding_payload(
    request: PreviewBankStatementImportRequest,
    source_sha256: str,
) -> dict[str, Any]:
    return {
        "version": "bank-statement-parser-v1",
        "org_id": request.org_id,
        "bank_account_code": request.bank_account_code,
        "file_format": request.file_format.value,
        "column_mapping": request.column_mapping.model_dump(exclude_none=True),
        "sheet_name": request.sheet_name,
        "date_format": request.date_format,
        "source_sha256": source_sha256,
    }


def _minimal_import_preview(
    source_sha256: str,
    *,
    errors: list[BankStatementIssue],
    rows: list[BankStatementPreviewRow] | None = None,
) -> BankStatementImportPreview:
    return BankStatementImportPreview(
        status=BankStatementPreviewStatus.REJECTED,
        source_sha256=source_sha256,
        rows=rows or [],
        errors=errors,
        trace=[
            {
                "stage": "bank_statement_preview_rejected",
                "error_codes": sorted({item.code for item in errors}),
            }
        ],
    )


def _validate_import_system_facts(
    request: PreviewBankStatementImportRequest,
    parsed_statement: ParsedBankStatement,
    system_facts: BankStatementImportSystemFacts,
) -> tuple[
    list[BankStatementIssue],
    dict[str, BankStatementImportRowSystemFact],
    dict[uuid.UUID, BankStatementImportEvidenceFact],
]:
    errors: list[BankStatementIssue] = []
    if (
        request.org_id != system_facts.org_id
        or request.bank_account_code != system_facts.bank_account_code
    ):
        errors.append(BankStatementIssue(code="BANK_STATEMENT_SYSTEM_FACTS_SCOPE_MISMATCH"))

    row_facts: dict[str, BankStatementImportRowSystemFact] = {}
    duplicate_row_fact = False
    for fact in system_facts.rows:
        if fact.row_identity_sha256 in row_facts:
            duplicate_row_fact = True
        row_facts[fact.row_identity_sha256] = fact
    expected_row_ids = {row.row_identity_sha256 for row in parsed_statement.rows}
    if duplicate_row_fact or set(row_facts) != expected_row_ids:
        errors.append(BankStatementIssue(code="BANK_STATEMENT_SYSTEM_ROW_FACTS_MISMATCH"))

    requested_evidence_ids = {
        evidence_id
        for resolution in request.missing_external_id_resolutions
        for evidence_id in resolution.evidence_references
    }
    evidence_facts: dict[uuid.UUID, BankStatementImportEvidenceFact] = {}
    duplicate_evidence_fact = False
    for fact in system_facts.resolution_evidence:
        if fact.evidence_id in evidence_facts:
            duplicate_evidence_fact = True
        evidence_facts[fact.evidence_id] = fact
        if fact.org_id != request.org_id:
            errors.append(
                BankStatementIssue(code="BANK_STATEMENT_RESOLUTION_EVIDENCE_SCOPE_MISMATCH")
            )
    if duplicate_evidence_fact or set(evidence_facts) != requested_evidence_ids:
        errors.append(BankStatementIssue(code="BANK_STATEMENT_RESOLUTION_EVIDENCE_FACTS_MISMATCH"))
    return errors, row_facts, evidence_facts


def _preview_row(
    row: BankStatementNormalizedRow,
    fact: BankStatementImportRowSystemFact,
) -> BankStatementPreviewRow:
    period = fact.period
    return BankStatementPreviewRow(
        **row.model_dump(),
        disposition=("ready" if row.external_id is not None else "needs_external_id_resolution"),
        period_status=period.status,
        period_id=period.period_id,
        is_late=period.status == "closed",
        original_close_id=period.close_id,
        original_close_hash=period.close_hash,
        original_closed_at=period.closed_at,
    )


def _apply_existing_external_id_snapshot(
    request: PreviewBankStatementImportRequest,
    row: BankStatementPreviewRow,
    fact: BankStatementImportRowSystemFact,
    errors: list[BankStatementIssue],
) -> None:
    existing = fact.existing_external_id_transaction
    if row.external_id is None:
        if existing is not None:
            errors.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_SYSTEM_ROW_FACTS_MISMATCH",
                    row_number=row.row_number,
                )
            )
        return
    if existing is None:
        return
    if existing.org_id != request.org_id or existing.bank_account_code != request.bank_account_code:
        errors.append(
            BankStatementIssue(
                code="BANK_STATEMENT_EXTERNAL_ID_TRANSACTION_SCOPE_MISMATCH",
                row_number=row.row_number,
            )
        )
        return
    if existing.external_id != row.external_id or not _same_transaction_facts(
        row, existing, include_descriptive_facts=True
    ):
        errors.append(
            BankStatementIssue(
                code="BANK_STATEMENT_EXTERNAL_ID_FACT_CONFLICT",
                row_number=row.row_number,
                field_path="external_id",
            )
        )
        return
    row.disposition = "stable_duplicate"
    row.duplicate_bank_transaction_id = existing.transaction_id


def _apply_existing_source_row_snapshot(
    request: PreviewBankStatementImportRequest,
    row: BankStatementPreviewRow,
    fact: BankStatementImportRowSystemFact,
    errors: list[BankStatementIssue],
) -> None:
    """Project an exact source-row replay before manual identity resolution.

    The database uniqueness guard is the concurrency backstop.  This projection
    gives a stable duplicate result for sequential replays that use a different
    idempotency key, without reviving the unsafe descriptive fingerprint rule.
    """

    existing = fact.existing_source_row_transaction
    if existing is None:
        return
    if (
        existing.org_id != request.org_id
        or existing.bank_account_code != request.bank_account_code
        or existing.external_id != row.external_id
        or not _same_transaction_facts(row, existing, include_descriptive_facts=True)
    ):
        # System facts come exclusively from the unique source-row lookup.  A
        # mismatch is an internal invariant breach and is handled by the caller
        # as a facts error rather than silently treating a different row as a
        # duplicate.
        errors.append(
            BankStatementIssue(
                code="BANK_STATEMENT_SOURCE_ROW_FACT_CONFLICT",
                row_number=row.row_number,
            )
        )
        return
    row.disposition = "stable_duplicate"
    row.duplicate_bank_transaction_id = existing.transaction_id


def _validate_manual_duplicate_snapshot(
    request: PreviewBankStatementImportRequest,
    row: BankStatementPreviewRow,
    fact: BankStatementImportRowSystemFact,
    resolution: MissingExternalIdResolution | None,
    errors: list[BankStatementIssue],
) -> None:
    target = fact.manual_duplicate_target
    expects_target = (
        resolution is not None
        and resolution.decision == "confirm_duplicate"
        and resolution.duplicate_bank_transaction_id is not None
    )
    if not expects_target:
        if target is not None:
            errors.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_SYSTEM_ROW_FACTS_MISMATCH",
                    row_number=row.row_number,
                )
            )
        return
    if target is None or target.transaction_id != resolution.duplicate_bank_transaction_id:
        errors.append(
            BankStatementIssue(
                code="BANK_STATEMENT_DUPLICATE_TARGET_FACTS_MISMATCH",
                row_number=row.row_number,
                field_path="duplicate_bank_transaction_id",
            )
        )
        return
    if target.org_id != request.org_id or target.bank_account_code != request.bank_account_code:
        errors.append(
            BankStatementIssue(
                code="BANK_STATEMENT_DUPLICATE_TARGET_SCOPE_MISMATCH",
                row_number=row.row_number,
                field_path="duplicate_bank_transaction_id",
            )
        )
        return
    if not _same_transaction_facts(row, target, include_descriptive_facts=False):
        errors.append(
            BankStatementIssue(
                code="BANK_STATEMENT_DUPLICATE_TARGET_FACT_CONFLICT",
                row_number=row.row_number,
                field_path="duplicate_bank_transaction_id",
            )
        )


def _same_transaction_facts(
    row: BankStatementPreviewRow,
    transaction: BankStatementTransactionSnapshot,
    *,
    include_descriptive_facts: bool,
) -> bool:
    core_matches = (
        row.booking_date == transaction.booking_date
        and row.amount_fen == transaction.amount_fen
        and row.currency == transaction.currency
    )
    if not include_descriptive_facts:
        return core_matches
    return (
        core_matches
        and row.counterparty_name == transaction.counterparty_name
        and row.memo == transaction.memo
    )


def _period_projection_matches_row(
    row: BankStatementPreviewRow,
    fact: BankStatementImportRowSystemFact,
) -> bool:
    expected_start = row.booking_date.replace(day=1)
    expected_end = row.booking_date.replace(
        day=monthrange(row.booking_date.year, row.booking_date.month)[1]
    )
    return (
        fact.period.period_start_date == expected_start
        and fact.period.period_end_date == expected_end
    )


def calculate_bank_reconciliation(
    request: PreviewBankReconciliationRequest,
    system_facts: BankReconciliationSystemFacts,
) -> BankReconciliationPreview:
    """Calculate DEC-014 B statement integrity and statement-to-book differences."""

    errors: list[BankStatementIssue] = []
    missing: list[BankStatementInformationRequirement] = []
    warnings: list[dict[str, object]] = []

    if (
        request.org_id != system_facts.org_id
        or request.period_id != system_facts.period_id
        or request.bank_account_code != system_facts.bank_account_code
    ):
        errors.append(BankStatementIssue(code="BANK_RECONCILIATION_SYSTEM_FACTS_MISMATCH"))

    missing_fields = [
        field_name
        for field_name in (
            "coverage_start_date",
            "coverage_end_date",
            "statement_opening_balance_fen",
            "statement_closing_balance_fen",
        )
        if getattr(request, field_name) is None
    ]
    if not request.statement_evidence_references:
        missing_fields.append("statement_evidence_references")
    if missing_fields:
        missing.append(
            BankStatementInformationRequirement(
                code="BANK_RECONCILIATION_INFORMATION_REQUIRED",
                fields=missing_fields,
            )
        )

    if len(request.statement_import_action_ids) != len(set(request.statement_import_action_ids)):
        errors.append(BankStatementIssue(code="BANK_RECONCILIATION_DUPLICATE_IMPORT_ACTION"))
    if len(request.statement_evidence_references) != len(
        set(request.statement_evidence_references)
    ):
        errors.append(BankStatementIssue(code="BANK_RECONCILIATION_DUPLICATE_EVIDENCE"))

    requested_action_ids = set(request.statement_import_action_ids)
    resolved_action_ids = {item.action_id for item in system_facts.import_actions}
    if (
        len(resolved_action_ids) != len(system_facts.import_actions)
        or requested_action_ids != resolved_action_ids
    ):
        errors.append(BankStatementIssue(code="BANK_RECONCILIATION_IMPORT_ACTION_FACTS_MISMATCH"))
    requested_evidence_ids = set(request.statement_evidence_references)
    resolved_evidence_ids = {item.evidence_id for item in system_facts.statement_evidence}
    if (
        len(resolved_evidence_ids) != len(system_facts.statement_evidence)
        or requested_evidence_ids != resolved_evidence_ids
    ):
        errors.append(BankStatementIssue(code="BANK_RECONCILIATION_EVIDENCE_FACTS_MISMATCH"))
    if any(
        item.org_id != request.org_id or item.bank_account_code != request.bank_account_code
        for item in system_facts.import_actions
    ):
        errors.append(BankStatementIssue(code="BANK_RECONCILIATION_IMPORT_ACTION_SCOPE_MISMATCH"))
    if any(item.org_id != request.org_id for item in system_facts.statement_evidence):
        errors.append(BankStatementIssue(code="BANK_RECONCILIATION_EVIDENCE_SCOPE_MISMATCH"))
    if system_facts.unmatched_transaction_count:
        warnings.append(
            {
                "code": "BANK_RECONCILIATION_UNMATCHED_TRANSACTIONS_REVIEW",
                "count": system_facts.unmatched_transaction_count,
            }
        )
    if system_facts.pending_late_transaction_count:
        warnings.append(
            {
                "code": "BANK_RECONCILIATION_PENDING_LATE_EVIDENCE_REVIEW",
                "count": system_facts.pending_late_transaction_count,
            }
        )

    transactions = [
        transaction for action in system_facts.import_actions for transaction in action.transactions
    ]
    transaction_ids = [item.transaction_id for item in transactions]
    if len(transaction_ids) != len(set(transaction_ids)):
        errors.append(BankStatementIssue(code="BANK_RECONCILIATION_DUPLICATE_TRANSACTION_FACT"))
    statement_transaction_count = len(transactions)
    statement_movement_fen = sum(item.amount_fen for item in transactions)
    if not _is_signed_bigint(statement_movement_fen):
        errors.append(BankStatementIssue(code="BANK_RECONCILIATION_AMOUNT_OUT_OF_RANGE"))

    coverage_start = request.coverage_start_date
    coverage_end = request.coverage_end_date
    if coverage_start is not None and coverage_end is not None:
        if coverage_start > coverage_end:
            errors.append(BankStatementIssue(code="BANK_RECONCILIATION_INVALID_COVERAGE"))
        elif (
            coverage_start != system_facts.period_start_date
            or coverage_end != system_facts.period_end_date
        ):
            missing.append(
                BankStatementInformationRequirement(
                    code="BANK_RECONCILIATION_COVERAGE_MUST_MATCH_PERIOD",
                    fields=["coverage_start_date", "coverage_end_date"],
                )
            )
        if any(
            item.booking_date < coverage_start or item.booking_date > coverage_end
            for item in transactions
        ):
            errors.append(
                BankStatementIssue(code="BANK_RECONCILIATION_TRANSACTION_OUTSIDE_COVERAGE")
            )

    opening = request.statement_opening_balance_fen
    closing = request.statement_closing_balance_fen
    statement_integrity_difference: int | None = None
    statement_to_book_difference: int | None = None
    if opening is not None and closing is not None:
        statement_integrity_difference = closing - opening - statement_movement_fen
        statement_to_book_difference = closing - system_facts.book_closing_balance_fen
        if not _is_signed_bigint(statement_integrity_difference) or not _is_signed_bigint(
            statement_to_book_difference
        ):
            errors.append(BankStatementIssue(code="BANK_RECONCILIATION_AMOUNT_OUT_OF_RANGE"))
        if statement_integrity_difference != 0:
            errors.append(
                BankStatementIssue(code="BANK_RECONCILIATION_STATEMENT_ROLLFORWARD_MISMATCH")
            )
        for kind, difference in {"statement_to_book": statement_to_book_difference}.items():
            explanations = [
                item for item in request.difference_explanations if item.difference_kind == kind
            ]
            if difference == 0:
                if explanations:
                    errors.append(
                        BankStatementIssue(
                            code="BANK_RECONCILIATION_UNEXPECTED_DIFFERENCE_EXPLANATION",
                            field_path=f"difference_explanations.{kind}",
                        )
                    )
                continue
            if not explanations:
                missing.append(
                    BankStatementInformationRequirement(
                        code="BANK_RECONCILIATION_DIFFERENCE_EXPLANATION_REQUIRED",
                        fields=[f"difference_explanations.{kind}"],
                    )
                )
                continue
            incomplete_fields: list[str] = []
            for index, explanation in enumerate(explanations):
                prefix = f"difference_explanations.{kind}.{index}"
                if explanation.amount_fen is None:
                    incomplete_fields.append(f"{prefix}.amount_fen")
                if explanation.explanation is None:
                    incomplete_fields.append(f"{prefix}.explanation")
                if not explanation.evidence_references:
                    incomplete_fields.append(f"{prefix}.evidence_references")
                elif not set(explanation.evidence_references) <= resolved_evidence_ids:
                    errors.append(
                        BankStatementIssue(
                            code=("BANK_RECONCILIATION_DIFFERENCE_EVIDENCE_FACTS_MISMATCH"),
                            field_path=f"{prefix}.evidence_references",
                        )
                    )
                if len(explanation.evidence_references) != len(
                    set(explanation.evidence_references)
                ):
                    errors.append(
                        BankStatementIssue(
                            code="BANK_RECONCILIATION_DUPLICATE_EVIDENCE",
                            field_path=f"{prefix}.evidence_references",
                        )
                    )
            if incomplete_fields:
                missing.append(
                    BankStatementInformationRequirement(
                        code="BANK_RECONCILIATION_DIFFERENCE_EXPLANATION_REQUIRED",
                        fields=incomplete_fields,
                    )
                )
                continue
            explained_total = sum(item.amount_fen or 0 for item in explanations)
            if explained_total != difference:
                errors.append(
                    BankStatementIssue(
                        code="BANK_RECONCILIATION_DIFFERENCE_EXPLANATION_MISMATCH",
                        field_path=f"difference_explanations.{kind}",
                    )
                )
            else:
                warnings.append(
                    {
                        "code": "BANK_RECONCILIATION_EXPLAINED_DIFFERENCE",
                        "difference_kind": kind,
                        "amount_fen": difference,
                    }
                )

    if errors:
        status = BankStatementPreviewStatus.REJECTED
    elif missing:
        status = BankStatementPreviewStatus.NEEDS_INFORMATION
    else:
        status = BankStatementPreviewStatus.CALCULATED

    calculation = {
        "version": "bank-reconciliation-v1",
        "org_id": request.org_id,
        "period_id": request.period_id,
        "bank_account_code": request.bank_account_code,
        "period_start_date": system_facts.period_start_date,
        "period_end_date": system_facts.period_end_date,
        "coverage_start_date": coverage_start,
        "coverage_end_date": coverage_end,
        "statement_opening_balance_fen": opening,
        "statement_closing_balance_fen": closing,
        "statement_transaction_count": statement_transaction_count,
        "statement_movement_fen": statement_movement_fen,
        "statement_integrity_difference_fen": statement_integrity_difference,
        "book_closing_balance_fen": system_facts.book_closing_balance_fen,
        "statement_to_book_difference_fen": statement_to_book_difference,
        "unmatched_transaction_count": system_facts.unmatched_transaction_count,
        "pending_late_transaction_count": system_facts.pending_late_transaction_count,
        "import_actions": sorted(
            (
                {
                    **item.model_dump(mode="json", exclude={"transactions"}),
                    "transactions": sorted(
                        (transaction.model_dump(mode="json") for transaction in item.transactions),
                        key=lambda transaction: str(transaction["transaction_id"]),
                    ),
                }
                for item in system_facts.import_actions
            ),
            key=lambda item: str(item["action_id"]),
        ),
        "statement_evidence": sorted(
            (item.model_dump(mode="json") for item in system_facts.statement_evidence),
            key=lambda item: str(item["evidence_id"]),
        ),
        "difference_explanations": sorted(
            (
                {
                    **item.model_dump(mode="json"),
                    "evidence_references": sorted(
                        str(evidence_id) for evidence_id in item.evidence_references
                    ),
                }
                for item in request.difference_explanations
            ),
            key=lambda item: (
                str(item["difference_kind"]),
                int(item["amount_fen"] or 0),
                str(item["explanation"] or ""),
                json.dumps(item["evidence_references"], sort_keys=True),
            ),
        ),
        "warnings": warnings,
    }
    calculation_hash = canonical_sha256(calculation) if status == "calculated" else None
    if status != BankStatementPreviewStatus.CALCULATED:
        return BankReconciliationPreview(
            status=status,
            missing_information=missing,
            errors=errors,
            warnings=[],
            trace=[
                {
                    "stage": (
                        "bank_reconciliation_rejected"
                        if status == BankStatementPreviewStatus.REJECTED
                        else "bank_reconciliation_needs_information"
                    )
                }
            ],
            data={},
        )
    return BankReconciliationPreview(
        status=status,
        calculation_hash=calculation_hash,
        missing_information=missing,
        errors=errors,
        warnings=warnings,
        trace=[
            {
                "stage": "bank_reconciliation_calculated",
                "statement_transaction_count": statement_transaction_count,
            },
            *(
                [{"stage": "calculation_hashed", "calculation_hash": calculation_hash}]
                if calculation_hash is not None
                else []
            ),
        ],
        data={"calculation": calculation},
    )


def _read_rows_once(
    request: PreviewBankStatementImportRequest,
    statement_bytes: bytes,
) -> list[tuple[int, dict[str, Any]]]:
    mapping = request.column_mapping.model_dump(exclude_none=True)
    if request.file_format is BankStatementFileFormat.CSV:
        rows = _read_csv(statement_bytes, mapping)
    elif request.file_format is BankStatementFileFormat.XLSX:
        rows = _read_xlsx(statement_bytes, request.sheet_name, mapping)
    else:  # pragma: no cover - schema fixes the finite format catalog.
        raise _InputError("BANK_STATEMENT_FORMAT_UNSUPPORTED")
    if not rows:
        raise _InputError("BANK_STATEMENT_EMPTY")
    return rows


def _read_csv(
    statement_bytes: bytes,
    mapping: dict[str, str],
) -> list[tuple[int, dict[str, Any]]]:
    decoded: str | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded = statement_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise _InputError("BANK_STATEMENT_PARSE_FAILED")
    try:
        reader = csv.DictReader(StringIO(decoded, newline=""))
        headers = reader.fieldnames or []
        _validate_source_columns(headers, mapping)
        result: list[tuple[int, dict[str, Any]]] = []
        for record_number, row in enumerate(reader, start=2):
            if record_number - 1 > MAX_STATEMENT_ROWS:
                raise _InputError("BANK_STATEMENT_FILE_LIMIT_EXCEEDED")
            if not any(_optional_text(value) is not None for value in row.values()):
                continue
            if None in row or any(value is None for value in row.values()):
                row = {"__bank_statement_structural_error__": True}
            _validate_source_cell_bounds(row.values())
            result.append((record_number, dict(row)))
        return result
    except _InputError:
        raise
    except (csv.Error, UnicodeError) as exc:
        raise _InputError("BANK_STATEMENT_PARSE_FAILED") from exc


def _read_xlsx(
    statement_bytes: bytes,
    sheet_name: str | None,
    mapping: dict[str, str],
) -> list[tuple[int, dict[str, Any]]]:
    _validate_xlsx_archive(statement_bytes)
    try:
        workbook = load_workbook(BytesIO(statement_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise _InputError("BANK_STATEMENT_PARSE_FAILED") from exc
    try:
        if sheet_name is not None and sheet_name not in workbook.sheetnames:
            raise _InputError("BANK_STATEMENT_SHEET_NOT_FOUND", field_path="sheet_name")
        worksheet = workbook[sheet_name] if sheet_name is not None else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            first = next(rows)
        except StopIteration as exc:
            raise _InputError("BANK_STATEMENT_EMPTY") from exc
        headers = [str(value).strip() if value is not None else "" for value in first]
        _validate_source_columns(headers, mapping)
        if worksheet.max_column > MAX_STATEMENT_COLUMNS:
            raise _InputError("BANK_STATEMENT_FILE_LIMIT_EXCEEDED")
        result: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(rows, start=2):
            if row_number - 1 > MAX_STATEMENT_ROWS:
                raise _InputError("BANK_STATEMENT_FILE_LIMIT_EXCEEDED")
            if not any(_optional_text(value) is not None for value in values):
                continue
            _validate_source_cell_bounds(values)
            result.append((row_number, dict(zip(headers, values, strict=True))))
        return result
    except _InputError:
        raise
    except Exception as exc:
        raise _InputError("BANK_STATEMENT_PARSE_FAILED") from exc
    finally:
        workbook.close()


def _validate_source_columns(headers: list[str], mapping: dict[str, str]) -> None:
    if not headers:
        raise _InputError("BANK_STATEMENT_EMPTY")
    if len(headers) > MAX_STATEMENT_COLUMNS:
        raise _InputError("BANK_STATEMENT_FILE_LIMIT_EXCEEDED")
    if len(headers) != len(set(headers)):
        raise _InputError("BANK_STATEMENT_DUPLICATE_SOURCE_COLUMN")
    header_names = set(headers)
    for canonical, source_name in mapping.items():
        if source_name not in header_names:
            raise _InputError(
                "BANK_STATEMENT_MISSING_COLUMN",
                field_path=f"column_mapping.{canonical}",
            )


def _validate_source_cell_bounds(values: Any) -> None:
    for value in values:
        if isinstance(value, str) and len(value) > MAX_SOURCE_CELL_CHARACTERS:
            raise _InputError("BANK_STATEMENT_FILE_LIMIT_EXCEEDED")


def _validate_xlsx_archive(statement_bytes: bytes) -> None:
    try:
        with zipfile.ZipFile(BytesIO(statement_bytes)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_MEMBERS:
                raise _InputError("BANK_STATEMENT_FILE_LIMIT_EXCEEDED")
            total = 0
            for member in members:
                normalized_name = member.filename.replace("\\", "/").casefold()
                if normalized_name.startswith("xl/externallinks/") or normalized_name.endswith(
                    "vbaproject.bin"
                ):
                    raise _InputError("BANK_STATEMENT_XLSX_ACTIVE_CONTENT_NOT_ALLOWED")
                total += member.file_size
                if (
                    member.file_size > MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES
                    or total > MAX_XLSX_TOTAL_UNCOMPRESSED_BYTES
                ):
                    raise _InputError("BANK_STATEMENT_FILE_LIMIT_EXCEEDED")
                compressed = max(member.compress_size, 1)
                if member.file_size > compressed * MAX_XLSX_COMPRESSION_RATIO:
                    raise _InputError("BANK_STATEMENT_FILE_LIMIT_EXCEEDED")
    except _InputError:
        raise
    except (zipfile.BadZipFile, OSError) as exc:
        raise _InputError("BANK_STATEMENT_PARSE_FAILED") from exc


def _normalize_row(
    row: dict[str, Any],
    mapping: dict[str, str],
    date_format: str | None,
) -> dict[str, Any]:
    if row.get("__bank_statement_structural_error__") is True:
        raise _RowError("BANK_STATEMENT_COLUMN_COUNT_MISMATCH", "row")

    def value(canonical: str, default: Any = None) -> Any:
        source_name = mapping.get(canonical)
        if source_name is None:
            return default
        if source_name not in row:
            raise _RowError("BANK_STATEMENT_MISSING_COLUMN", canonical)
        return row[source_name]

    try:
        booking_date = _parse_date(value("booking_date"), date_format)
    except (TypeError, ValueError, OverflowError) as exc:
        raise _RowError("BANK_STATEMENT_INVALID_DATE", "booking_date") from exc
    if "amount" in mapping:
        try:
            amount = _decimal(value("amount"))
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise _RowError("BANK_STATEMENT_INVALID_AMOUNT", "amount") from exc
        amount_field = "amount"
        amount_fen = _decimal_to_fen(amount, amount_field)
    else:
        try:
            credit = _decimal(value("credit", 0), blank_zero=True)
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise _RowError("BANK_STATEMENT_INVALID_AMOUNT", "credit") from exc
        try:
            debit = _decimal(value("debit", 0), blank_zero=True)
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise _RowError("BANK_STATEMENT_INVALID_AMOUNT", "debit") from exc
        credit_fen = _decimal_to_fen(credit, "credit")
        debit_fen = _decimal_to_fen(debit, "debit")
        amount_fen = credit_fen - debit_fen
        amount_field = "credit"
    if amount_fen == 0:
        raise _RowError("BANK_STATEMENT_ZERO_AMOUNT", amount_field)
    if not _is_signed_bigint(amount_fen):
        raise _RowError("BANK_STATEMENT_AMOUNT_OUT_OF_RANGE", amount_field)
    currency = str(value("currency", "CNY") or "CNY").strip().upper()
    if currency != "CNY":
        raise _RowError("BANK_STATEMENT_INVALID_CURRENCY", "currency")
    return {
        "external_id": _bounded_optional_text(
            value("external_id"), max_length=100, field_path="external_id"
        ),
        "booking_date": booking_date,
        "amount_fen": amount_fen,
        "currency": currency,
        "counterparty_name": _bounded_optional_text(
            value("counterparty"), max_length=200, field_path="counterparty"
        ),
        "memo": _bounded_optional_text(value("memo"), max_length=2000, field_path="memo") or "",
    }


def _apply_external_id_resolutions(
    resolutions: list[MissingExternalIdResolution],
    rows_by_identity: dict[str, BankStatementPreviewRow],
    missing: list[BankStatementInformationRequirement],
) -> list[BankStatementIssue]:
    issues: list[BankStatementIssue] = []
    seen: set[str] = set()
    for resolution in resolutions:
        identity = resolution.row_identity_sha256
        if identity in seen:
            issues.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_DUPLICATE_EXTERNAL_ID_RESOLUTION",
                    row_number=resolution.row_number,
                    field_path="missing_external_id_resolutions",
                )
            )
            continue
        seen.add(identity)
        row = rows_by_identity.get(identity)
        if row is None or row.row_number != resolution.row_number:
            issues.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_EXTERNAL_ID_RESOLUTION_STALE",
                    row_number=resolution.row_number,
                    field_path="missing_external_id_resolutions.row_identity_sha256",
                )
            )
            continue
        if row.external_id is None and row.disposition == "stable_duplicate":
            # The exact file row has already been confirmed.  Preserve replay
            # semantics even when the original explicit resolution is supplied
            # again under a new idempotency key.
            continue
        if row.external_id is not None:
            issues.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_EXTERNAL_ID_RESOLUTION_NOT_ALLOWED",
                    row_number=row.row_number,
                    field_path="missing_external_id_resolutions",
                )
            )
            continue
        fields: list[str] = []
        prefix = f"missing_external_id_resolutions.{identity}"
        if resolution.decision is None:
            fields.append(f"{prefix}.decision")
        if resolution.explanation is None:
            fields.append(f"{prefix}.explanation")
        if not resolution.evidence_references:
            fields.append(f"{prefix}.evidence_references")
        if len(resolution.evidence_references) != len(set(resolution.evidence_references)):
            issues.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_DUPLICATE_EVIDENCE_REFERENCE",
                    row_number=row.row_number,
                    field_path=f"{prefix}.evidence_references",
                )
            )
        if resolution.decision == "confirm_duplicate":
            if resolution.duplicate_bank_transaction_id is None:
                fields.append(f"{prefix}.duplicate_bank_transaction_id")
        elif resolution.duplicate_bank_transaction_id is not None:
            issues.append(
                BankStatementIssue(
                    code="BANK_STATEMENT_DUPLICATE_TARGET_NOT_ALLOWED",
                    row_number=row.row_number,
                    field_path=f"{prefix}.duplicate_bank_transaction_id",
                )
            )
        if fields:
            missing.append(
                BankStatementInformationRequirement(
                    code="BANK_STATEMENT_EXTERNAL_ID_RESOLUTION_REQUIRED",
                    fields=fields,
                )
            )
            continue
        if resolution.decision == "confirm_new":
            row.disposition = "manual_new"
        elif resolution.decision == "confirm_duplicate":
            row.disposition = "manual_duplicate"
            row.duplicate_bank_transaction_id = resolution.duplicate_bank_transaction_id
    return issues


def _parse_date(value: Any, date_format: str | None) -> date:
    if isinstance(value, bool):
        raise ValueError("boolean is not a date")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if date_format:
        return datetime.strptime(text, date_format).date()
    return date.fromisoformat(text)


def _decimal(value: Any, *, blank_zero: bool = False) -> Decimal:
    if isinstance(value, bool):
        raise ValueError("boolean is not an amount")
    if isinstance(value, float):
        raise ValueError("binary floating point is not an exact monetary fact")
    if value is None or str(value).strip() == "":
        if blank_zero:
            return Decimal(0)
        raise ValueError("amount is blank")
    cleaned = str(value).strip().replace(",", "").replace("¥", "").replace("￥", "")
    if len(cleaned) > 100:
        raise InvalidOperation
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    result = Decimal(cleaned)
    if not result.is_finite():
        raise InvalidOperation
    return result


def _decimal_to_fen(value: Decimal, field_path: str) -> int:
    decimal_tuple = value.as_tuple()
    digits = decimal_tuple.digits
    exponent = int(decimal_tuple.exponent)
    coefficient_text = "".join(str(digit) for digit in digits) or "0"
    significant_text = coefficient_text.lstrip("0")
    if not significant_text:
        return 0
    trailing_zero_count = len(coefficient_text) - len(coefficient_text.rstrip("0"))
    effective_exponent = exponent + trailing_zero_count
    if effective_exponent < -2:
        raise _RowError("BANK_STATEMENT_AMOUNT_NOT_EXACT_FEN", field_path)
    scaled_exponent = exponent + 2
    # Reject impossible BIGINT values before exponentiation.  Decimal accepts
    # arbitrarily large exponents, so computing ``10**scaled_exponent`` first
    # would make a tiny statement cell a CPU/memory denial of service.
    projected_fen_digits = len(significant_text) + scaled_exponent
    if projected_fen_digits > 19:
        raise _RowError("BANK_STATEMENT_AMOUNT_OUT_OF_RANGE", field_path)
    coefficient = int(coefficient_text)
    if scaled_exponent >= 0:
        amount_fen = coefficient * (10**scaled_exponent)
    else:
        divisor = 10 ** (-scaled_exponent)
        amount_fen, remainder = divmod(coefficient, divisor)
        if remainder:
            raise _RowError("BANK_STATEMENT_AMOUNT_NOT_EXACT_FEN", field_path)
    if decimal_tuple.sign:
        amount_fen = -amount_fen
    if not _is_signed_bigint(amount_fen):
        raise _RowError("BANK_STATEMENT_AMOUNT_OUT_OF_RANGE", field_path)
    return amount_fen


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _bounded_optional_text(
    value: Any,
    *,
    max_length: int,
    field_path: str,
) -> str | None:
    text = _optional_text(value)
    if text is not None and len(text) > max_length:
        raise _RowError("BANK_STATEMENT_TEXT_TOO_LONG", field_path)
    return text


def _is_signed_bigint(value: int) -> bool:
    return MIN_SIGNED_BIGINT <= value <= MAX_SIGNED_BIGINT


def _canonical_value(value: Any) -> Any:
    if isinstance(value, bool) or value is None or isinstance(value, (str, int)):
        return value
    if isinstance(value, float):
        raise TypeError("binary floating point is not canonical accounting input")
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _canonical_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_canonical_value(item) for item in value]
    raise TypeError(f"unsupported canonical value type: {type(value).__name__}")
