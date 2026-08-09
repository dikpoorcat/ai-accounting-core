from __future__ import annotations

import csv
import hashlib
import json
import uuid
from collections.abc import Iterable
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import BankTransaction, Organization
from .schemas import ImportBankStatementRequest

CANONICAL_COLUMNS = {
    "booking_date",
    "amount",
    "debit",
    "credit",
    "counterparty",
    "memo",
    "external_id",
    "currency",
}


class BankStatementInputError(ValueError):
    """A stable, caller-safe failure while opening or parsing a statement."""

    def __init__(self, code: str, *, field: str | None = None) -> None:
        super().__init__(code)
        self.code = code
        self.field = field


class BankStatementRowError(ValueError):
    """A stable, caller-safe validation failure for one input record."""

    def __init__(self, code: str, field: str) -> None:
        super().__init__(code)
        self.code = code
        self.field = field


def import_bank_statement(session: Session, request: ImportBankStatementRequest) -> dict[str, Any]:
    if session.get(Organization, request.org_id) is None:
        raise ValueError("ORGANIZATION_NOT_FOUND")
    try:
        path = request.file_path.resolve(strict=True)
    except OSError as exc:
        raise BankStatementInputError("BANK_STATEMENT_FILE_UNAVAILABLE") from exc
    extension = path.suffix.lower()
    if extension not in {".csv", ".xlsx"}:
        raise BankStatementInputError("BANK_STATEMENT_FORMAT_UNSUPPORTED")
    mapping = request.column_mapping.model_dump(exclude_none=True)
    _validate_mapping(mapping)
    source_sha = hashlib.sha256(path.read_bytes()).hexdigest()
    try:
        rows = (
            _read_csv(path, mapping)
            if extension == ".csv"
            else _read_xlsx(path, request.sheet_name, mapping)
        )
    except BankStatementInputError:
        raise
    except Exception as exc:
        raise BankStatementInputError("BANK_STATEMENT_PARSE_FAILED") from exc

    imported: list[str] = []
    duplicates: list[str] = []
    errors: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        try:
            normalized = _normalize_row(row, mapping, request.date_format)
            fingerprint = _fingerprint(
                request.org_id,
                request.bank_account_code,
                normalized,
            )
            existing = session.scalar(
                select(BankTransaction.id).where(
                    BankTransaction.org_id == request.org_id,
                    BankTransaction.fingerprint == fingerprint,
                )
            )
            if existing:
                duplicates.append(str(existing))
                continue
            transaction = BankTransaction(
                org_id=request.org_id,
                bank_account_code=request.bank_account_code,
                fingerprint=fingerprint,
                source_sha256=source_sha,
                **normalized,
            )
            session.add(transaction)
            session.flush()
            imported.append(str(transaction.id))
        except BankStatementRowError as exc:
            errors.append({"row": row_number, "field": exc.field, "code": exc.code})
    return {
        "source_sha256": source_sha,
        "imported_count": len(imported),
        "duplicate_count": len(duplicates),
        "error_count": len(errors),
        "imported_ids": imported,
        "duplicate_ids": duplicates,
        "errors": errors,
    }


def _validate_mapping(mapping: dict[str, str]) -> None:
    unknown = set(mapping) - CANONICAL_COLUMNS
    if unknown:
        raise ValueError(f"unknown canonical columns: {sorted(unknown)}")
    if "booking_date" not in mapping:
        raise ValueError("column_mapping must include booking_date")
    if "amount" not in mapping and not ({"debit", "credit"} <= set(mapping)):
        raise ValueError("map amount, or map both debit and credit")


def _read_csv(path: Path, mapping: dict[str, str]) -> Iterable[dict[str, Any]]:
    raw = path.read_bytes()
    decoded: str | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise BankStatementInputError("BANK_STATEMENT_PARSE_FAILED")
    reader = csv.DictReader(decoded.splitlines())
    _validate_source_columns(reader.fieldnames or [], mapping)
    return list(reader)


def _read_xlsx(
    path: Path, sheet_name: str | None, mapping: dict[str, str]
) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        _validate_source_columns(headers, mapping)
        return [dict(zip(headers, values, strict=True)) for values in rows]
    finally:
        workbook.close()


def _validate_source_columns(headers: Iterable[str], mapping: dict[str, str]) -> None:
    header_names = set(headers)
    for canonical, source_name in mapping.items():
        if source_name not in header_names:
            raise BankStatementInputError(
                "BANK_STATEMENT_MISSING_COLUMN", field=canonical
            )


def _normalize_row(
    row: dict[str, Any], mapping: dict[str, str], date_format: str | None
) -> dict[str, Any]:
    def value(canonical: str, default: Any = None) -> Any:
        source_name = mapping.get(canonical)
        if source_name is None:
            return default
        if source_name not in row:
            raise BankStatementRowError("BANK_STATEMENT_MISSING_COLUMN", canonical)
        return row[source_name]

    try:
        booking_date = _parse_date(value("booking_date"), date_format)
    except (TypeError, ValueError, OverflowError) as exc:
        raise BankStatementRowError("BANK_STATEMENT_INVALID_DATE", "booking_date") from exc
    if "amount" in mapping:
        try:
            amount = _decimal(value("amount"))
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise BankStatementRowError("BANK_STATEMENT_INVALID_AMOUNT", "amount") from exc
    else:
        try:
            credit = _decimal(value("credit", 0), blank_zero=True)
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise BankStatementRowError("BANK_STATEMENT_INVALID_AMOUNT", "credit") from exc
        try:
            debit = _decimal(value("debit", 0), blank_zero=True)
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise BankStatementRowError("BANK_STATEMENT_INVALID_AMOUNT", "debit") from exc
        amount = credit - debit
    try:
        amount_fen = int((amount * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, OverflowError, ValueError) as exc:
        raise BankStatementRowError(
            "BANK_STATEMENT_INVALID_AMOUNT", "amount" if "amount" in mapping else "credit"
        ) from exc
    if amount_fen == 0:
        raise BankStatementRowError(
            "BANK_STATEMENT_ZERO_AMOUNT", "amount" if "amount" in mapping else "credit"
        )
    currency = str(value("currency", "CNY") or "CNY").strip().upper()
    if currency != "CNY":
        raise BankStatementRowError("BANK_STATEMENT_INVALID_CURRENCY", "currency")
    return {
        "external_id": _optional_text(value("external_id")),
        "booking_date": booking_date,
        "amount_fen": amount_fen,
        "currency": currency,
        "counterparty_name": _optional_text(value("counterparty")),
        "memo": _optional_text(value("memo")) or "",
    }


def _parse_date(value: Any, date_format: str | None) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if date_format:
        return datetime.strptime(text, date_format).date()
    return date.fromisoformat(text)


def _decimal(value: Any, *, blank_zero: bool = False) -> Decimal:
    if value is None or str(value).strip() == "":
        if blank_zero:
            return Decimal(0)
        raise ValueError("amount is blank")
    cleaned = str(value).strip().replace(",", "").replace("¥", "").replace("￥", "")
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    return Decimal(cleaned)


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _fingerprint(org_id: uuid.UUID, bank_account_code: str, normalized: dict[str, Any]) -> str:
    identity = {
        "org_id": str(org_id),
        "bank_account_code": bank_account_code,
        "external_id": normalized["external_id"],
        "booking_date": normalized["booking_date"].isoformat(),
        "amount_fen": normalized["amount_fen"],
        "counterparty_name": normalized["counterparty_name"],
        "memo": normalized["memo"],
    }
    return hashlib.sha256(
        json.dumps(identity, sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()
