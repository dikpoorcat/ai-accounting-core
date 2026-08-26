from __future__ import annotations

import hashlib
import io
import re
import zipfile
from datetime import UTC, date, datetime

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ai_accounting.accounting_period_schemas import PreviewAccountingPeriodCloseRequest
from ai_accounting.accounting_period_service import AccountingPeriodService
from ai_accounting.coa import seed_organization
from ai_accounting.financial_statement_schemas import (
    ConfirmEnterpriseIncomeTaxQuarterRequest,
    ConfirmFinancialStatementClassificationRequest,
    EnterpriseIncomeTaxTreatment,
    FinancialStatementResultStatus,
    PreviewQuarterlyFinancialStatementsRequest,
)
from ai_accounting.financial_statements import (
    TEMPLATE_SHA256,
    FinancialStatementService,
    _template_bytes,
)
from ai_accounting.ledger import Entry, create_voucher
from ai_accounting.models import (
    Account,
    AccountingPeriod,
    AccountingPeriodAction,
    AccountingPeriodCalendar,
    AccountingPeriodClose,
    BusinessEvent,
    Counterparty,
    Evidence,
    FinancialStatementClassification,
    Organization,
    Voucher,
    VoucherLine,
)
from ai_accounting.schemas import ReverseEventRequest
from ai_accounting.service import FinanceService


def _evidence(session: Session, organization: Organization, name: str) -> Evidence:
    digest = hashlib.sha256(name.encode()).hexdigest()
    row = Evidence(
        org_id=organization.id,
        sha256=digest,
        original_name=name,
        media_type="text/plain",
        source="test",
        size_bytes=1,
        storage_path=f"tests/{name}",
        metadata_json={},
    )
    session.add(row)
    session.flush()
    return row


def _open_quarter(session: Session, organization: Organization) -> list[AccountingPeriod]:
    organization.accounting_period_control_enabled = True
    organization.accounting_period_control_start_date = date(2026, 1, 1)
    calendar = AccountingPeriodCalendar(
        org_id=organization.id,
        calendar_year=2026,
        rule_version="test",
        rule_effective_from=date(2026, 1, 1),
        source_urls=["https://example.test/calendar"],
    )
    session.add(calendar)
    session.flush()
    periods: list[AccountingPeriod] = []
    ends = {1: 31, 2: 28, 3: 31}
    for month in range(1, 4):
        action = AccountingPeriodAction(
            org_id=organization.id,
            action_type="period_generation",
            idempotency_key=f"generate-2026-{month:02d}",
            request_payload_hash=hashlib.sha256(str(month).encode()).hexdigest(),
            status="posted",
            input_facts={"month": month},
            missing_information=[],
            errors=[],
            confirmed_by="test",
            confirmation_note="测试生成期间",
        )
        session.add(action)
        session.flush()
        period = AccountingPeriod(
            org_id=organization.id,
            calendar_id=calendar.id,
            generation_action_id=action.id,
            calendar_year=2026,
            calendar_month=month,
            start_date=date(2026, month, 1),
            end_date=date(2026, month, ends[month]),
            status="open",
        )
        session.add(period)
        session.flush()
        periods.append(period)
    return periods


def _close_quarter(
    session: Session, organization: Organization, periods: list[AccountingPeriod]
) -> None:
    previous_hash = None
    for period in periods:
        action = AccountingPeriodAction(
            org_id=organization.id,
            action_type="period_close",
            idempotency_key=f"close-2026-{period.calendar_month:02d}",
            request_payload_hash=hashlib.sha256(
                f"close-{period.calendar_month}".encode()
            ).hexdigest(),
            status="posted",
            input_facts={"month": period.calendar_month},
            missing_information=[],
            errors=[],
            confirmed_by="test",
            confirmation_note="测试关闭期间",
        )
        session.add(action)
        session.flush()
        calculation_hash = hashlib.sha256(
            f"close-snapshot-{period.calendar_month}".encode()
        ).hexdigest()
        close = AccountingPeriodClose(
            org_id=organization.id,
            period_id=period.id,
            action_id=action.id,
            calculation={},
            calculation_payload="{}",
            calculation_hash=calculation_hash,
            rule_version="test",
            rule_effective_from=date(2026, 1, 1),
            source_urls=["https://example.test/close"],
            previous_close_hash=previous_hash,
            checker_version="test",
            confirmed_at=datetime.now(UTC),
            voucher_count=0,
            line_count=0,
            total_debit_fen=0,
            total_credit_fen=0,
        )
        session.add(close)
        session.flush()
        period.status = "closed"
        period.closed_at = datetime.now(UTC)
        period.close_id = close.id
        previous_hash = calculation_hash
    session.flush()


def _post(
    session: Session,
    organization: Organization,
    *,
    key: str,
    event_type: str,
    posting_date: date,
    entries: list[Entry],
) -> BusinessEvent:
    event = BusinessEvent(
        org_id=organization.id,
        idempotency_key=key,
        request_payload_hash=hashlib.sha256(key.encode()).hexdigest(),
        event_type=event_type,
        status="draft",
        description=key,
        facts={},
        business_date=posting_date,
        posting_date=posting_date,
        rule_trace=[],
        rule_version="test",
    )
    session.add(event)
    session.flush()
    create_voucher(
        session,
        event=event,
        posting_date=posting_date,
        description=key,
        entries=entries,
    )
    event.status = "posted"
    session.flush()
    return event


def _prepare_calculated_q1(
    session: Session, organization: Organization
) -> tuple[FinancialStatementService, PreviewQuarterlyFinancialStatementsRequest]:
    periods = _open_quarter(session, organization)
    evidence = _evidence(session, organization, "quarterly-report.txt")
    _post(
        session,
        organization,
        key="capital",
        event_type="owner_contribution_received",
        posting_date=date(2026, 1, 2),
        entries=[
            Entry(account_role="bank", debit_fen=100_000),
            Entry(account_role="paid_in_capital", credit_fen=100_000),
        ],
    )
    _post(
        session,
        organization,
        key="sale",
        event_type="service_cash_sale",
        posting_date=date(2026, 2, 3),
        entries=[
            Entry(account_role="bank", debit_fen=50_000),
            Entry(account_role="service_revenue", credit_fen=50_000),
        ],
    )
    expense = _post(
        session,
        organization,
        key="expense",
        event_type="expense_cash",
        posting_date=date(2026, 3, 4),
        entries=[
            Entry(account_role="general_expense", debit_fen=10_000),
            Entry(account_role="bank", credit_fen=10_000),
        ],
    )
    _post(
        session,
        organization,
        key="fixed-asset-acquisition",
        event_type="fixed_asset_acquisition",
        posting_date=date(2026, 3, 5),
        entries=[
            Entry(account_role="fixed_asset_cost", debit_fen=70_000),
            Entry(account_role="bank", credit_fen=70_000),
        ],
    )
    expense_line = session.scalar(
        select(VoucherLine)
        .join(Voucher, Voucher.id == VoucherLine.voucher_id)
        .where(Voucher.event_id == expense.id)
        .where(VoucherLine.debit_fen == 10_000)
    )
    assert expense_line is not None
    service = FinancialStatementService(session)
    classification = service.confirm_classification(
        ConfirmFinancialStatementClassificationRequest(
            org_id=organization.id,
            voucher_line_id=expense_line.id,
            allocations=[{"detail_code": "management_other", "amount_fen": 10_000}],
            idempotency_key="classify-expense",
            confirmation_note="确认全部为其他管理费用",
            evidence_references=[evidence.id],
        )
    )
    assert classification.status is FinancialStatementResultStatus.POSTED
    _close_quarter(session, organization, periods)
    income_tax = service.confirm_enterprise_income_tax(
        ConfirmEnterpriseIncomeTaxQuarterRequest(
            org_id=organization.id,
            year=2026,
            quarter=1,
            treatment=EnterpriseIncomeTaxTreatment.ZERO,
            amount_fen=0,
            idempotency_key="cit-q1-zero",
            confirmation_note="本季度明确确认所得税费用为零",
            evidence_references=[evidence.id],
        )
    )
    assert income_tax.status is FinancialStatementResultStatus.POSTED
    return service, PreviewQuarterlyFinancialStatementsRequest(
        org_id=organization.id, year=2026, quarter=1
    )


def test_requirements_block_unclassified_expense_and_missing_income_tax(
    session: Session, organization: Organization
) -> None:
    periods = _open_quarter(session, organization)
    _post(
        session,
        organization,
        key="unclassified-expense",
        event_type="expense_cash",
        posting_date=date(2026, 1, 5),
        entries=[
            Entry(account_role="general_expense", debit_fen=1_000),
            Entry(account_role="bank", credit_fen=1_000),
        ],
    )
    _close_quarter(session, organization, periods)

    result = FinancialStatementService(session).preview_quarterly(
        PreviewQuarterlyFinancialStatementsRequest(org_id=organization.id, year=2026, quarter=1)
    )

    assert result.status is FinancialStatementResultStatus.NEEDS_INFORMATION
    codes = {item.code for item in result.missing_information}
    assert "FINANCIAL_STATEMENT_CLASSIFICATION_REQUIRED" in codes
    assert "ENTERPRISE_INCOME_TAX_QUARTER_CONFIRMATION_REQUIRED" in codes


def test_quarter_end_close_requires_income_tax_confirmation(
    session: Session, organization: Organization
) -> None:
    periods = _open_quarter(session, organization)

    result = AccountingPeriodService(
        session, current_date=date.max
    ).preview_accounting_period_close(
        PreviewAccountingPeriodCloseRequest(
            org_id=organization.id,
            period_id=periods[2].id,
            closing_date=periods[2].end_date,
        )
    )

    assert "ACCOUNTING_PERIOD_ENTERPRISE_INCOME_TAX_CONFIRMED" in result.data["blocker_codes"]


def test_quarterly_statements_are_deterministic_and_balanced(
    session: Session, organization: Organization
) -> None:
    service, request = _prepare_calculated_q1(session, organization)

    first = service.preview_quarterly(request)
    second = service.preview_quarterly(request)

    assert first.status is FinancialStatementResultStatus.CALCULATED
    assert first.calculation_hash == second.calculation_hash
    statements = first.data["statements"]
    assert statements["balance_sheet"]["30"]["ending_fen"] == 140_000
    assert statements["balance_sheet"]["53"]["ending_fen"] == 140_000
    assert statements["balance_sheet"]["20"]["ending_fen"] == 70_000
    assert statements["profit_statement"]["1"]["current_fen"] == 50_000
    assert statements["profit_statement"]["14"]["current_fen"] == 10_000
    assert statements["profit_statement"]["32"]["current_fen"] == 40_000
    assert statements["cash_flow_statement"]["12"]["current_fen"] == 70_000
    assert statements["cash_flow_statement"]["20"]["current_fen"] == 70_000
    assert statements["cash_flow_statement"]["22"]["current_fen"] == 70_000
    assert all(item["passed"] for item in first.data["checks"])


def test_balance_sheet_reclassifies_receivables_and_payables_by_counterparty(
    session: Session, organization: Organization
) -> None:
    customer_receivable = Counterparty(
        org_id=organization.id,
        kind="customer",
        name="应收客户",
    )
    customer_advance = Counterparty(
        org_id=organization.id,
        kind="customer",
        name="预收客户",
    )
    supplier_prepayment = Counterparty(
        org_id=organization.id,
        kind="supplier",
        name="预付供应商",
    )
    supplier_payable = Counterparty(
        org_id=organization.id,
        kind="supplier",
        name="应付供应商",
    )
    session.add_all([customer_receivable, customer_advance, supplier_prepayment, supplier_payable])
    session.flush()
    postings = [
        (
            "ar-debit",
            Entry(
                account_role="accounts_receivable",
                debit_fen=1_000,
                counterparty_id=customer_receivable.id,
            ),
            Entry(account_role="paid_in_capital", credit_fen=1_000),
        ),
        (
            "ar-credit",
            Entry(
                account_role="accounts_receivable",
                credit_fen=500,
                counterparty_id=customer_advance.id,
            ),
            Entry(account_role="paid_in_capital", debit_fen=500),
        ),
        (
            "ap-debit",
            Entry(
                account_role="accounts_payable",
                debit_fen=800,
                counterparty_id=supplier_prepayment.id,
            ),
            Entry(account_role="paid_in_capital", credit_fen=800),
        ),
        (
            "ap-credit",
            Entry(
                account_role="accounts_payable",
                credit_fen=300,
                counterparty_id=supplier_payable.id,
            ),
            Entry(account_role="paid_in_capital", debit_fen=300),
        ),
    ]
    for key, first_entry, second_entry in postings:
        _post(
            session,
            organization,
            key=key,
            event_type="balance_reclassification_test",
            posting_date=date(2026, 1, 5),
            entries=[first_entry, second_entry],
        )

    service = FinancialStatementService(session)
    missing = []
    balance = service._balance_sheet(
        service._ledger_rows(organization.id, date(2026, 3, 31)),
        date(2026, 1, 1),
        date(2026, 3, 31),
        missing,
    )

    assert not missing
    assert balance[4]["ending_fen"] == 1_000
    assert balance[34]["ending_fen"] == 500
    assert balance[5]["ending_fen"] == 800
    assert balance[33]["ending_fen"] == 300
    assert balance[30]["ending_fen"] == balance[53]["ending_fen"] == 1_800


def test_cash_flow_reversal_uses_original_event_mapping(
    session: Session, organization: Organization
) -> None:
    periods = _open_quarter(session, organization)
    evidence = _evidence(session, organization, "cash-flow-reversal.txt")
    source = _post(
        session,
        organization,
        key="cash-sale-to-reverse",
        event_type="service_cash_sale",
        posting_date=date(2026, 1, 5),
        entries=[
            Entry(account_role="bank", debit_fen=1_000),
            Entry(account_role="service_revenue", credit_fen=1_000),
        ],
    )
    reversal = FinanceService(session).reverse_event(
        ReverseEventRequest(
            org_id=organization.id,
            event_id=source.id,
            idempotency_key="reverse-cash-sale",
            reason="测试季度内销售退款冲正",
            posting_date=date(2026, 2, 5),
        )
    )
    assert reversal.status == "posted"
    _close_quarter(session, organization, periods)
    service = FinancialStatementService(session)
    income_tax = service.confirm_enterprise_income_tax(
        ConfirmEnterpriseIncomeTaxQuarterRequest(
            org_id=organization.id,
            year=2026,
            quarter=1,
            treatment=EnterpriseIncomeTaxTreatment.ZERO,
            amount_fen=0,
            idempotency_key="reversal-income-tax-zero",
            confirmation_note="明确确认本季度所得税费用为零",
            evidence_references=[evidence.id],
        )
    )
    assert income_tax.status == "posted"

    result = service.preview_quarterly(
        PreviewQuarterlyFinancialStatementsRequest(
            org_id=organization.id,
            year=2026,
            quarter=1,
        )
    )

    assert result.status is FinancialStatementResultStatus.CALCULATED
    assert result.data["statements"]["profit_statement"]["1"]["current_fen"] == 0
    assert result.data["statements"]["cash_flow_statement"]["1"]["current_fen"] == 0
    assert result.data["statements"]["cash_flow_statement"]["22"]["current_fen"] == 0


def test_unmapped_nonzero_cash_event_blocks_export(
    session: Session, organization: Organization
) -> None:
    periods = _open_quarter(session, organization)
    evidence = _evidence(session, organization, "unmapped-cash-event.txt")
    _post(
        session,
        organization,
        key="unmapped-cash-event",
        event_type="future_cash_event",
        posting_date=date(2026, 1, 5),
        entries=[
            Entry(account_role="bank", debit_fen=500),
            Entry(account_role="paid_in_capital", credit_fen=500),
        ],
    )
    _close_quarter(session, organization, periods)
    service = FinancialStatementService(session)
    income_tax = service.confirm_enterprise_income_tax(
        ConfirmEnterpriseIncomeTaxQuarterRequest(
            org_id=organization.id,
            year=2026,
            quarter=1,
            treatment=EnterpriseIncomeTaxTreatment.ZERO,
            amount_fen=0,
            idempotency_key="unmapped-cash-income-tax-zero",
            confirmation_note="明确确认本季度所得税费用为零",
            evidence_references=[evidence.id],
        )
    )
    assert income_tax.status == "posted"

    result, workbook = service.export_quarterly_xlsx(
        PreviewQuarterlyFinancialStatementsRequest(
            org_id=organization.id,
            year=2026,
            quarter=1,
        )
    )

    assert result.status is FinancialStatementResultStatus.NEEDS_INFORMATION
    assert workbook is None
    assert "FINANCIAL_STATEMENT_UNMAPPED_CASH_EVENT" in {
        item.code for item in result.missing_information
    }


def test_classification_is_append_only_and_idempotent(
    session: Session, organization: Organization
) -> None:
    _open_quarter(session, organization)
    evidence = _evidence(session, organization, "classification.txt")
    event = _post(
        session,
        organization,
        key="classifiable-expense",
        event_type="expense_cash",
        posting_date=date(2026, 1, 5),
        entries=[
            Entry(account_role="general_expense", debit_fen=1_000),
            Entry(account_role="bank", credit_fen=1_000),
        ],
    )
    line = session.scalar(
        select(VoucherLine)
        .join(Voucher, Voucher.id == VoucherLine.voucher_id)
        .where(Voucher.event_id == event.id)
        .where(VoucherLine.debit_fen == 1_000)
    )
    assert line is not None
    service = FinancialStatementService(session)
    request = ConfirmFinancialStatementClassificationRequest(
        org_id=organization.id,
        voucher_line_id=line.id,
        allocations=[{"detail_code": "management_other", "amount_fen": 1_000}],
        idempotency_key="classification-v1",
        confirmation_note="首次分类",
        evidence_references=[evidence.id],
    )
    first = service.confirm_classification(request)
    replay = service.confirm_classification(request)
    assert first.status is FinancialStatementResultStatus.POSTED
    assert replay.classification_id == first.classification_id
    assert replay.data["idempotent_replay"] is True

    correction = service.confirm_classification(
        ConfirmFinancialStatementClassificationRequest(
            org_id=organization.id,
            voucher_line_id=line.id,
            allocations=[{"detail_code": "management_startup", "amount_fen": 1_000}],
            supersedes_classification_id=first.classification_id,
            idempotency_key="classification-v2",
            confirmation_note="更正为开办费",
            evidence_references=[evidence.id],
        )
    )
    assert correction.status is FinancialStatementResultStatus.POSTED
    assert (
        session.scalar(
            select(FinancialStatementClassification).where(
                FinancialStatementClassification.id == first.classification_id
            )
        )
        is not None
    )
    original = session.get(FinancialStatementClassification, first.classification_id)
    assert original is not None
    original.confirmation_note = "禁止修改"
    with pytest.raises(ValueError, match="FINANCIAL_STATEMENT_FACT_IMMUTABLE"):
        session.flush()
    session.expire(original)


def test_income_tax_accrual_and_reduction_are_controlled_entries(
    session: Session, organization: Organization
) -> None:
    _open_quarter(session, organization)
    evidence = _evidence(session, organization, "income-tax.txt")
    service = FinancialStatementService(session)
    accrued = service.confirm_enterprise_income_tax(
        ConfirmEnterpriseIncomeTaxQuarterRequest(
            org_id=organization.id,
            year=2026,
            quarter=1,
            treatment=EnterpriseIncomeTaxTreatment.ACCRUE,
            amount_fen=2_000,
            posting_date=date(2026, 3, 31),
            idempotency_key="cit-accrue",
            confirmation_note="确认本季度所得税费用",
            evidence_references=[evidence.id],
        )
    )
    assert accrued.status is FinancialStatementResultStatus.POSTED
    assert accrued.event_id is not None
    lines = list(
        session.scalars(select(VoucherLine).where(VoucherLine.voucher_id == accrued.voucher_id))
    )
    assert sorted((line.debit_fen, line.credit_fen) for line in lines) == [
        (0, 2_000),
        (2_000, 0),
    ]

    reduction_org = seed_organization(
        session,
        name="所得税冲减测试公司",
        taxpayer_identification_number="91330106MA1234567T",
    )
    _open_quarter(session, reduction_org)
    reduction_evidence = _evidence(session, reduction_org, "income-tax-reduction.txt")
    _post(
        session,
        reduction_org,
        key="opening-cit-accrual",
        event_type="enterprise_income_tax_assessment",
        posting_date=date(2026, 1, 31),
        entries=[
            Entry(account_role="enterprise_income_tax_expense", debit_fen=3_000),
            Entry(account_role="enterprise_income_tax_payable", credit_fen=3_000),
        ],
    )
    reduced = FinancialStatementService(session).confirm_enterprise_income_tax(
        ConfirmEnterpriseIncomeTaxQuarterRequest(
            org_id=reduction_org.id,
            year=2026,
            quarter=1,
            treatment=EnterpriseIncomeTaxTreatment.REDUCE,
            amount_fen=2_000,
            posting_date=date(2026, 3, 31),
            idempotency_key="cit-reduce",
            confirmation_note="冲减前期多计所得税",
            evidence_references=[reduction_evidence.id],
        )
    )
    assert reduced.status is FinancialStatementResultStatus.POSTED
    reduction_lines = list(
        session.scalars(select(VoucherLine).where(VoucherLine.voucher_id == reduced.voucher_id))
    )
    roles_by_id = {
        account.id: account.system_role
        for account in session.scalars(select(Account).where(Account.org_id == reduction_org.id))
    }
    assert {
        (roles_by_id[line.account_id], line.debit_fen, line.credit_fen) for line in reduction_lines
    } == {
        ("enterprise_income_tax_payable", 2_000, 0),
        ("enterprise_income_tax_expense", 0, 2_000),
    }


def test_tax_template_preserves_structure_and_cached_values(
    session: Session, organization: Organization
) -> None:
    service, request = _prepare_calculated_q1(session, organization)
    result, generated = service.export_quarterly_xlsx(request)
    assert result.status is FinancialStatementResultStatus.CALCULATED
    assert generated is not None
    template = _template_bytes()
    assert hashlib.sha256(template).hexdigest().upper() == TEMPLATE_SHA256

    source = load_workbook(io.BytesIO(template), data_only=False)
    output = load_workbook(io.BytesIO(generated), data_only=False)
    cached = load_workbook(io.BytesIO(generated), data_only=True)
    assert (
        output.sheetnames
        == source.sheetnames
        == ["资产负债表", "利润表_月季报", "现金流量表_月季报"]
    )
    for source_sheet, output_sheet in zip(source.worksheets, output.worksheets, strict=True):
        assert source_sheet.protection.sheet == output_sheet.protection.sheet
        assert tuple(source_sheet.merged_cells.ranges) == tuple(output_sheet.merged_cells.ranges)
        assert len(source_sheet.data_validations.dataValidation) == len(
            output_sheet.data_validations.dataValidation
        )
        for row in source_sheet.iter_rows():
            for source_cell in row:
                if source_cell.data_type == "f":
                    assert output_sheet[source_cell.coordinate].value == source_cell.value
    assert cached["资产负债表"]["D37"].value == 1400
    assert cached["资产负债表"]["I37"].value == 0
    assert cached["利润表_月季报"]["D3"].value == "91330106MA1234567T"
    assert cached["利润表_月季报"]["F3"].value == organization.name
    assert cached["利润表_月季报"]["D4"].value.date() == date(2026, 1, 1)
    assert cached["现金流量表_月季报"]["F4"].value.date() == date(2026, 3, 31)
    assert cached["利润表_月季报"]["D37"].value == 400
    assert cached["现金流量表_月季报"]["D30"].value == 700

    with zipfile.ZipFile(io.BytesIO(template)) as source_zip, zipfile.ZipFile(
        io.BytesIO(generated)
    ) as output_zip:
        assert source_zip.namelist() == output_zip.namelist()
        for part_name, root_name in (
            ("xl/worksheets/sheet1.xml", "worksheet"),
            ("xl/worksheets/sheet2.xml", "worksheet"),
            ("xl/worksheets/sheet3.xml", "worksheet"),
            ("xl/workbook.xml", "workbook"),
        ):
            source_xml = source_zip.read(part_name).decode("utf-8")
            output_xml = output_zip.read(part_name).decode("utf-8")
            root_pattern = rf"<{root_name}\b[^>]*>"
            source_root = re.search(root_pattern, source_xml)
            output_root = re.search(root_pattern, output_xml)
            assert source_root is not None and output_root is not None
            assert output_root.group(0) == source_root.group(0)
            ignorable = re.search(r'mc:Ignorable="([^"]+)"', output_root.group(0))
            if ignorable is not None:
                for prefix in ignorable.group(1).split():
                    assert f"xmlns:{prefix}=" in output_root.group(0)
